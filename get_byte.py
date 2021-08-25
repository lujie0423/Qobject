# -*- coding: utf-8 -*-
import os
import os.path
import psutil
import openpyxl
import time, datetime
import pythoncom
import uuid
from PIL import ImageGrab, Image
from openpyxl.reader.excel import load_workbook
from win32com.client import Dispatch, DispatchEx
import xlwings as xw
import test_excel as f



android_path="//10.21.100.89/Project Arukas/Versions/android/SBT/"
ios_path="//10.21.100.89/Project Arukas/Versions/ios/SBT/"

lists_android = os.listdir(android_path)
# print(lists_android)
lists_ios = os.listdir(ios_path)
# print(lists_ios)

lists_android.sort(key=lambda x:os.path.getmtime((android_path+"\\"+x)))
lists_ios.sort(key=lambda x:os.path.getmtime((ios_path+"\\"+x)))
# print("时间排序后的的文件夹列表：\n %s \n"%lists_android )
# print("时间排序后的的文件夹列表：\n %s \n"%lists_ios )




# 判断当天是否有重复的打包
#把最新修改的20个文件，天数放入list中
day_list = []
for i in range(20):
    day_list.append(time.localtime(os.stat(os.path.join(android_path, lists_android[-i-1])).st_mtime).tm_mday)
# print (day_list)
# 如果最新修改有多个，代表一天内打了多次包，算出个数
len_today = 0
for i in range(len(day_list)):
    if day_list[0] == day_list[i]:
        len_today = len_today + 1

file_android = os.path.join(android_path, lists_android[-1])
file_ios = os.path.join(ios_path, lists_ios[-1])
# print("今日文件路径:\n%s"%file_android)
# print("今日文件路径:\n%s"%file_ios)

yes_file_android = os.path.join(android_path, lists_android[-len_today-1])
yes_file_ios = os.path.join(ios_path, lists_ios[-len_today-1])
# print("昨日文件路径:\n%s"%yes_file_android)
# print("昨日文件路径:\n%s"%yes_file_ios)

android = lists_android[-1]
ios = lists_ios[-1]
print("今日最新文件:\n%s" % android)
print("今日最新文件:\n%s" % ios)

yes_android = lists_android[-len_today-1]
yes_ios = lists_ios[-len_today-1]
print("昨日最新文件:\n%s" % yes_android)
print("昨日最新文件:\n%s" % yes_ios)

# 文件最后修改时间 - 时间戳
end_change_android = os.stat(file_android).st_mtime
end_change_ios = os.stat(file_ios).st_mtime
# print(end_change_android)
# print(end_change_android)
yes_end_change_android = os.stat(yes_file_android).st_mtime
yes_end_change_ios = os.stat(yes_file_ios).st_mtime

# 时间戳转换
now_android = time.localtime(end_change_android)
now_ios = time.localtime(end_change_ios)
now_time = datetime.datetime.now()
# print(now_android.tm_mday)
# print(now_ios.tm_mday)

yes_now_android = time.localtime(yes_end_change_android)
yes_now_ios = time.localtime(yes_end_change_ios)
oneday = datetime.timedelta(days=1)
yes_now_time = now_time - oneday
# print(yes_time)

if now_android.tm_mday == now_time.day == now_ios.tm_mday:
    print("\033[0;31;40m今日打包成功\033[0m")

elif now_android.tm_mday != now_time.day:
    print("\033[0;31;40m今日android打包失败\033[0m")
elif now_ios.tm_mday != now_time.day:
    print("\033[0;31;40m今日ios打包失败\033[0m")
else:
    print("打包失败")



if yes_now_android.tm_mday == yes_now_time.day == yes_now_ios.tm_mday:
    print("昨天打包成功")
elif yes_now_android.tm_mday != yes_now_time.day:
    print("昨天ios打包失败")
elif yes_now_ios.tm_mday != yes_now_time.day:
    print("昨天android打包失败")
else:
    print("打包失败")

# print(yes_now_android.tm_mday)
# print(yes_now_time.day)
# print(yes_now_ios.tm_mday)

#今日
apk_file = file_android + '/' +  android +'.apk'
print(apk_file)
ipa_file = file_ios + '/' + ios + '_enterprise.ipa'
print(ipa_file)


#获得文件大小，并得到MB
size_android = '%.2f' % (os.path.getsize(apk_file) / (1024**2)) + 'MB'
size_ios = '%.2f' % (os.path.getsize(ipa_file) / (1024**2)) + 'MB'
print("今日包android：",size_android)
print("今日包ios：",size_ios)


#昨日
yes_apk_file = yes_file_android + '/' +  yes_android +'.apk'
print(yes_apk_file)
yes_ipa_file = yes_file_ios + '/' + yes_ios + '_enterprise.ipa'
print(yes_ipa_file)


#获得昨日文件大小，并得到MB
size_yes_android = '%.2f' % (os.path.getsize(yes_apk_file) / (1024**2)) + 'MB'
size_yes_ios = '%.2f' % (os.path.getsize(yes_ipa_file) / (1024**2)) + 'MB'
print("昨日包android：",size_yes_android)
print("昨日包ios：",size_yes_ios)


# # excel文件绝对路径
file_home = "C:/Users/jie.lu/Desktop/SMOKE_版本.xlsx"
wb = load_workbook(filename=file_home)  # 打开excel文件
sheet_ranges = wb['TOTAL']
# print(sheet_ranges['A1'].value)  # 打印A1单元格的值
ws = wb['TOTAL']  # 根据Sheet1这个sheet名字来获取该sheet
wd = wb['ARU']
# 修改android & ios的版本号
ws['c5'] = android
ws['c6'] = ios
# 修改今日版本大小
ws['i5'] = size_android
ws['i6'] = size_ios
# 修改昨日版本大小
ws['j5'] = size_yes_android
ws['j6'] = size_yes_ios

# 修改ARU - sheet的版本
wd['c6'] = android + "\n" + ios
#获取内容，拆分字符串
str = wd['b12'].value
len_str = len(str)
#字符串拆分
now_month = now_time.strftime('%m')
now_day = now_time.strftime('%d')
now = now_month + "月" + now_day + "号"
head = str[:6]
middle = now
end = str[-(len_str - 12):]
all = head + middle + end
wd['b12'] = all
wb.save(file_home)  # 保存修改后的excel
print("修改excel成功")


# 打开excel并截图保存
# os.startfile(file_home)
# xb = xw.Book(file_home)
# # app = xw.App(add_book=False)
# # app.books.open(file_home)
# sht = xw.Sheet(sheet='TOTAL')
# sht.select()
# time.sleep(2)
# app.kill()
#
#
# file_pic1 = "C:\\Users\\jie.lu\\Desktop\\每日报告\\1.png"
# file_pic2 = "C:\\Users\\jie.lu\\Desktop\\每日报告\\2.png"
# # 参数说明
# # 第一个参数 开始截图的x坐标
# # 第二个参数 开始截图的y坐标
# # 第三个参数 结束截图的x坐标
# # 第四个参数 结束截图的y坐标
#
# #228 正文与顶部距离
# #26  正文与侧部距离
#
#
# bbox = (126, 227, 1475, 543)
# im = ImageGrab.grab(bbox)
# # 参数 保存截图文件的路径
# im.save(file_pic1)
#
# cbox = (126, 578, 430, 671)
# im = ImageGrab.grab(cbox)
# im.save(file_pic2)


f.excel_catch_screen(file_home, "TOTAL", "B1:J6", "C:\\Users\jie.lu\\Desktop\\每日报告\\1.png")
f.excel_catch_screen(file_home, "TOTAL", "B9:C13", "C:\\Users\jie.lu\\Desktop\\每日报告\\2.png")

#打开图片看看是否截取成功
# os.startfile(file_pic1)
# os.startfile(file_pic2)

